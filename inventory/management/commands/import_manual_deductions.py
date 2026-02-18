import csv
import json
import re
import unicodedata
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from inventory.models import InventoryProduct, InventoryRecipeItem, InventoryVariation, StockBucket


def _norm(value):
    value = str(value or "").strip().lower().replace("\u0131", "i")
    value = unicodedata.normalize("NFKD", value)
    value = value.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", value)


def _canonical_header(raw):
    key = _norm(raw)
    if key in {"urun", "product"}:
        return "product"
    if key in {"variation", "varyasyon"}:
        return "variation"
    if key in {"stockbucket", "stockbuckets", "bucket"}:
        return "bucket"
    if key in {"dusulecekmiktar", "miktar", "quantity", "qty"}:
        return "quantity"
    return ""


def _read_json(path):
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise CommandError("JSON root must be a list of rows.")
    rows = []
    for row in data:
        if not isinstance(row, dict):
            continue
        mapped = {}
        for key, value in row.items():
            canonical = _canonical_header(key)
            if canonical:
                mapped[canonical] = value
        rows.append(mapped)
    return rows


def _read_csv(path):
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise CommandError("CSV header row is missing.")
        mapped = {_canonical_header(h): h for h in reader.fieldnames if _canonical_header(h)}
        required = {"product", "variation", "bucket", "quantity"}
        if not required.issubset(mapped.keys()):
            raise CommandError(
                "CSV must include columns for product, variation, bucket, quantity."
            )
        for row in reader:
            rows.append(
                {
                    "product": row.get(mapped["product"]),
                    "variation": row.get(mapped["variation"]),
                    "bucket": row.get(mapped["bucket"]),
                    "quantity": row.get(mapped["quantity"]),
                }
            )
    return rows


def _read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CommandError(
            "openpyxl is required for .xlsx import. Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    values = ws.iter_rows(values_only=True)
    try:
        header = next(values)
    except StopIteration:
        return []

    header_map = {}
    for idx, raw in enumerate(header):
        canonical = _canonical_header(raw)
        if canonical:
            header_map[canonical] = idx
    required = {"product", "variation", "bucket", "quantity"}
    if not required.issubset(header_map.keys()):
        raise CommandError(
            "XLSX must include columns for product, variation, bucket, quantity."
        )

    rows = []
    for row in values:
        rows.append(
            {
                "product": row[header_map["product"]] if len(row) > header_map["product"] else "",
                "variation": row[header_map["variation"]] if len(row) > header_map["variation"] else "",
                "bucket": row[header_map["bucket"]] if len(row) > header_map["bucket"] else "",
                "quantity": row[header_map["quantity"]] if len(row) > header_map["quantity"] else "",
            }
        )
    return rows


def _read_rows(path):
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _read_json(path)
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise CommandError("Unsupported file type. Use .json, .csv, or .xlsx")


class Command(BaseCommand):
    help = "Import manual variation->bucket deduction rules into InventoryRecipeItem."

    def add_arguments(self, parser):
        parser.add_argument("--user-id", type=int, required=True)
        parser.add_argument("--file", type=str, required=True)
        parser.add_argument(
            "--clear-existing",
            action="store_true",
            help="Delete all existing recipe rows for this user's products before import.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and print summary without writing DB changes.",
        )

    def handle(self, *args, **options):
        user_id = options["user_id"]
        file_path = Path(options["file"])
        clear_existing = options["clear_existing"]
        dry_run = options["dry_run"]

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        User = get_user_model()
        try:
            owner = User.objects.get(id=user_id)
        except User.DoesNotExist as exc:
            raise CommandError(f"User not found: {user_id}") from exc

        rows = _read_rows(file_path)
        if not rows:
            self.stdout.write(self.style.WARNING("No rows found."))
            return

        products = list(InventoryProduct.objects.filter(owner=owner))
        products_by_norm = {_norm(p.name): p for p in products}

        updates = []
        errors = []
        for i, row in enumerate(rows, start=2):
            product_name = str(row.get("product") or "").strip()
            variation_name = str(row.get("variation") or "").strip()
            bucket_name = str(row.get("bucket") or "").strip()
            quantity_raw = row.get("quantity")

            if not (product_name and variation_name and bucket_name):
                errors.append(f"Row {i}: missing product/variation/bucket value.")
                continue
            try:
                quantity = int(quantity_raw)
            except (TypeError, ValueError):
                errors.append(f"Row {i}: invalid quantity '{quantity_raw}'.")
                continue
            if quantity <= 0:
                errors.append(f"Row {i}: quantity must be > 0.")
                continue

            product = products_by_norm.get(_norm(product_name))
            if not product:
                errors.append(f"Row {i}: product not found: '{product_name}'.")
                continue

            variations = list(InventoryVariation.objects.filter(product=product))
            variation_by_norm = {_norm(v.name): v for v in variations}
            variation = variation_by_norm.get(_norm(variation_name))
            if not variation:
                errors.append(
                    f"Row {i}: variation not found for product '{product.name}': '{variation_name}'."
                )
                continue

            buckets = list(
                StockBucket.objects.filter(owner=owner, product=product, is_active=True)
            )
            bucket_by_norm = {_norm(b.name): b for b in buckets}
            bucket = bucket_by_norm.get(_norm(bucket_name))
            if not bucket:
                errors.append(
                    f"Row {i}: bucket not found for product '{product.name}': '{bucket_name}'."
                )
                continue

            updates.append((variation, bucket, quantity))

        if errors:
            preview = "\n".join(errors[:20])
            more = "" if len(errors) <= 20 else f"\n... and {len(errors) - 20} more"
            raise CommandError(f"Import failed with {len(errors)} error(s):\n{preview}{more}")

        created = 0
        updated = 0
        deleted = 0

        if dry_run:
            self.stdout.write(
                self.style.SUCCESS(
                    f"Dry run ok. Parsed {len(rows)} rows, ready to apply {len(updates)} recipe rows."
                )
            )
            return

        with transaction.atomic():
            if clear_existing:
                deleted, _ = InventoryRecipeItem.objects.filter(
                    variation__product__owner=owner
                ).delete()

            for variation, bucket, quantity in updates:
                obj, was_created = InventoryRecipeItem.objects.update_or_create(
                    variation=variation,
                    bucket=bucket,
                    defaults={"quantity": quantity},
                )
                if was_created:
                    created += 1
                elif obj.quantity == quantity:
                    updated += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. rows={len(rows)} applied={len(updates)} created={created} updated={updated} deleted={deleted}"
            )
        )
